# tests/test_sonarcloud_export.py
import unittest
from unittest.mock import patch, MagicMock
from utils import filter_issues_by_component, get_open_issues
from constants import EXCLUSION_PATTERNS, ISSUE_FIELDS, SONARCLOUD_API_URL, OUTPUT_FILE
import openpyxl
import os

from sonarcloud_export import write_issues_to_excel

class TestSonarCloudExport(unittest.TestCase):

    def test_filter_issues_by_component(self):
        issues = [
            {"component": "included_component"},
            {"component": EXCLUSION_PATTERNS[0]}
        ]
        filtered_issues = filter_issues_by_component(issues)
        self.assertEqual(len(filtered_issues), 1)
        self.assertEqual(filtered_issues[0]["component"], "included_component")

    @patch("utils.requests.get")
    def test_get_open_issues(self, mock_get):
        mock_response = MagicMock()
        mock_response.json.return_value = {
            "issues": [
                {
                    "message": "Test message",
                    "type": "BUG",
                    "impacts": [{"severity": "CRITICAL"}],
                    "effort": "5min",
                    "component": "test_component",
                    "line": 10
                }
            ],
            "paging": {"total": 1}
        }
        mock_response.status_code = 200
        mock_get.return_value = mock_response

        token = "test_token"
        project_key = "test_project_key"
        issues = get_open_issues(project_key, token)
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0]["message"], "Test message")

    def test_write_issues_to_excel(self):
        project_keys = ["test_project_key"]
        token = "test_token"
        output_file = "test_output.xlsx"

        with patch("utils.get_open_issues") as mock_get_open_issues:
            mock_get_open_issues.return_value = [
                {
                    "message": "Test message",
                    "type": "BUG",
                    "severity": "CRITICAL",
                    "effort": "5min",
                    "component": "test_component",
                    "line": 10
                }
            ]

            write_issues_to_excel(project_keys, token, output_file)

            # Verify the output file was created
            self.assertTrue(os.path.exists(output_file))

            # Load the workbook and verify the contents
            workbook = openpyxl.load_workbook(output_file)
            sheet = workbook.active
            self.assertEqual(sheet.title, "test_project_key")
            self.assertEqual(sheet.cell(row=1, column=1).value, "Message")
            self.assertEqual(sheet.cell(row=2, column=1).value, "Test message")

            # Clean up
            os.remove(output_file)

if __name__ == "__main__":
    unittest.main()
